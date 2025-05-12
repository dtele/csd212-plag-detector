import os
import sys
import logging
import warnings
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple

import openpyxl as xl
import pandas as pd
import yaml
from box import Box

import copydetect.detector as cd


logging.basicConfig(level=logging.INFO)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


@dataclass
class DetectorFilePathConfig:
    demo: int
    preprocessed_assignments_folder: str
    plagiarism_excel_save_path: str
    teams_excel_path: str
    demo_folder: str = None

    def __post_init__(self):
        self.demo_folder = os.path.join(self.preprocessed_assignments_folder, f'Demo{self.demo}')

    @classmethod
    def from_dict(cls, d: dict):
        return cls(**d)


@dataclass
class FingerprintConfig:
    k: int
    win_size: int
    boilerplate: List[str]

    @classmethod
    def from_dict(cls, d: dict):
        return cls(**d)


@dataclass
class PlagiarismConfig:
    penalties: list[tuple[float, float]]

    @classmethod
    def from_dict(cls, d: dict):
        penalties = list(sorted(
            (float(threshold), float(penalty))
            for threshold, penalty in d['penalties'].items()
        ))
        return cls(penalties=penalties)


@dataclass
class ExcelConfig:
    teams_row_start: int
    teams_name_col: str
    teams_marks_col: str
    plagiarism_row_start =  2
    plagiarism_name_col = 'A'
    plagiarism_similarity_col = 'B'
    plagiarism_bool_col = 'D'
    plagiarism_marks_col = 'E'
    plagiarism_final_marks_col = 'F'

    @classmethod
    def from_dict(cls, d: dict):
        return cls(**d)


@dataclass
class DetectorConfig:
    paths: DetectorFilePathConfig
    fingerprint: FingerprintConfig
    plagiarism: PlagiarismConfig
    excel: ExcelConfig

    @classmethod
    def from_dict(cls, d: dict):
        return cls(
            paths=DetectorFilePathConfig.from_dict(d['paths']),
            fingerprint=FingerprintConfig.from_dict(d['fingerprint']),
            plagiarism=PlagiarismConfig.from_dict(d['plagiarism']),
            excel=ExcelConfig.from_dict(d['excel']),
        )


def generate_file_fingerprints(submission_path: str, k: int, win_size: int, boilerplate: Optional[List[str]] = None) -> Dict[str, cd.CodeFingerprint]:
    """
    Generates fingerprints for all files in a folder.

    :param submission_path: Path to the submission folder of a single assignment.
    :param k: Length of k-grams to extract as fingerprints.
    :param win_size: Window size for winnowing algorithm.
    :param boilerplate: Optional list of common strings (import statements, plt.show etc. to ignore from plagiarism check).
    :return: Dictionary of filename and its fingerprint.
    """
    try:
        file_fingerprints = {}

        for file in os.listdir(submission_path):
            file_path = os.path.join(submission_path, file)
            file_fingerprints[os.path.splitext(file)[0]] = cd.CodeFingerprint(file_path, k, win_size, boilerplate)

        logging.info(f'generate_file_fingerprints()\n'
              f'Successfully generated fingerprints for files at `{submission_path}`.\n')

        return file_fingerprints
    except FileNotFoundError:
        logging.error(f'generate_file_fingerprints()\n'
              f'Error in generating file fingerprints for files at `{submission_path}`.\n'
              'Kindly ensure that the folder exists and has been preprocessed.\n')
        sys.exit(1)


def find_max_similarities(file_fingerprints: Dict[str, cd.CodeFingerprint]) -> Dict[str, Tuple[float, str]]:
    """
    Returns the maximum similarity score for each file in ``file_fingerprints`` and the file it's with.

    :param file_fingerprints: Dictionary of filename and its fingerprint for files to be checked.
    :return: Dictionary of filename and the maximum similarity found.
    """
    try:
        max_similarities = {}
        checked_pairs = set()

        for file1, fingerprint1 in file_fingerprints.items():
            checked_pairs.add((file1, file1))
            max_similarity = (0, '')

            for file2, fingerprint2 in file_fingerprints.items():
                if (file1, file2) in checked_pairs or (file2, file1) in checked_pairs:
                    continue

                token_overlap, similarities, slices = cd.compare_files(fingerprint1, fingerprint2)

                if similarities and similarities[0] > max_similarity[0]:
                    max_similarity = float(similarities[0]), file2

            max_similarities[file1] = max_similarity

        logging.info(f'find_max_similarities()\n'
              'Successfully compared file fingerprints.\n')

        return max_similarities
    except Exception as e:
        logging.error(f'find_max_similarities()\n'
              'Error in comparing file fingerprints. Kindly verify files manually.\n')
        sys.exit(1)


def create_plagiarism_excel(save_path: str, max_similarities: Dict[str, Tuple[float, str]], plagiarism_config: PlagiarismConfig, excel_config: ExcelConfig) -> None:
    """
    Creates an Excel file with the columns

    ``student_name`` `: name,` ``max_similarity`` `: float`, ``similarity_with`` `: name,` ``plagiarism_penalized`` `: bool,` ``marks_given`` `: user input,` ``final_marks`` `: penalized marks_given based on threshold`

    :param save_path: Path to where the Excel file will be saved.
    :param max_similarities: Dictionary of filename and the maximum similarity found from ``find_max_similarities()``.
    :param plagiarism_config: PlagiarismConfig object containing a list of thresholds and their corresponding penalties.
    :param excel_config: ExcelConfig object containing starting row and column alphabet configuration.
    :return: None
    """
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    try:
        rows = []

        for student_name, (max_similarity, similarity_with) in max_similarities.items():
            rows.append({
                'student_name': student_name,
                'max_similarity': max_similarity,
                'similarity_with': similarity_with,
                'plagiarism_penalized': None,
                'marks_given': None,
                'final_marks': None
            })

        df = pd.DataFrame(rows)
        df.to_excel(save_path, index=False)

        wb = xl.load_workbook(save_path)
        ws = wb.active

        for row in range(excel_config.plagiarism_row_start, ws.max_row + 1):
            max_similarity_cell = f'{excel_config.plagiarism_similarity_col}{row}'
            plagiarism_cell = f'{excel_config.plagiarism_bool_col}{row}'
            given_marks_cell = f'{excel_config.plagiarism_marks_col}{row}'
            final_marks_cell = f'{excel_config.plagiarism_final_marks_col}{row}'

            ws[plagiarism_cell] = f'=IF({max_similarity_cell} >= {plagiarism_config.penalties[0][0]}, True, False)'

            conditions = []
            for threshold, penalty in reversed(plagiarism_config.penalties):
                conditions.append(f'IF({max_similarity_cell} >= {threshold}, {given_marks_cell} * (1 - {penalty})')

            ws[final_marks_cell] = '=' + ', '.join(conditions) + f', {given_marks_cell}'+ (')' * len(conditions))

        for i, column in enumerate(df.columns):
            letter = chr(65 + i)

            max_length = max(max(map(len, df['student_name'].astype(str))), len(column))
            ws.column_dimensions[letter].width = max_length + 2

        wb.save(save_path)
        logging.info('create_plagiarism_excel()\n'
              f'Created plagiarism adjusted Excel at `{save_path}` successfully.\n')
    except PermissionError:
        logging.error('create_plagiarism_excel()\n'
              f'Error creating plagiarism adjusted Excel. Kindly check that excel workbook at `{save_path}` is not currently open.\n')
        sys.exit(1)


def merge_teams_excel(teams_excel_path: str, plagiarism_excel_path: str, excel_config: ExcelConfig) -> None:
    """
    Adds `marks_given` to plagiarism adjusted Excel from the Excel downloaded from teams.

    :param teams_excel_path: Path to the Excel downloaded from teams.
    :param plagiarism_excel_path: Path to the Excel generated with ``create_plagiarism_excel()``.
    :param excel_config: ExcelConfig object containing starting row and column alphabet configuration.
    :return: None
    """
    try:
        teams_wb = xl.load_workbook(teams_excel_path)
        plagiarism_wb = xl.load_workbook(plagiarism_excel_path)
    except PermissionError:
        logging.error('merge_teams_excel()\n'
              f'Error in merging grades from teams excel. '
              f'Kindly check that excel workbook at `{plagiarism_excel_path}` is not currently open.\n')
        sys.exit(1)

    teams_ws = teams_wb.active
    plagiarism_ws = plagiarism_wb.active

    missed_students = set()
    marks_given = {}
    for row in range(excel_config.teams_row_start, teams_ws.max_row + 1):
        name_cell = f'{excel_config.teams_name_col}{row}'
        marks_cell = f'{excel_config.teams_marks_col}{row}'
        name = teams_ws[name_cell].value
        marks = teams_ws[marks_cell].value

        name = ''.join([char for char in name if char.isalpha() or char == ' ']).strip()

        if name:
            marks_given[name] = marks
            missed_students.add(name or None)

    all_students_count = len(missed_students)
    for row in range(excel_config.plagiarism_row_start, plagiarism_ws.max_row + 1):
        name_cell = f'{excel_config.plagiarism_name_col}{row}'
        marks_cell = f'{excel_config.plagiarism_marks_col}{row}'
        name = plagiarism_ws[name_cell].value

        try:
            plagiarism_ws[marks_cell] = marks_given[name]
            missed_students.remove(name)
        except KeyError:
            pass

    if not missed_students:
        logging.info('merge_teams_excel()\n'
              f'Entries for {all_students_count - len(missed_students)}/{all_students_count} students updated successfully.\n')
    else:
        missed_students = sorted(list(missed_students))
        logging.info('merge_teams_excel()\n'
              f'Entries for {all_students_count - len(missed_students)}/{all_students_count} students updated successfully.')
        logging.warning(f'Kindly update plagiarism for following students manually: {", ".join(missed_students)}\n')

        for name in missed_students:
            plagiarism_ws.append([name, 0, '', False, marks_given[name], marks_given[name]])

    plagiarism_wb.save(plagiarism_excel_path)


with open('config.yaml', 'r') as f:
    config = Box(yaml.safe_load(f))

detector_config = DetectorConfig.from_dict(config.detector)

paths = detector_config.paths
fingerprint_config = detector_config.fingerprint
plagiarism_config = detector_config.plagiarism
excel_config = detector_config.excel

if __name__ == '__main__':
    fps = generate_file_fingerprints(paths.demo_folder, fingerprint_config.k, fingerprint_config.win_size, fingerprint_config.boilerplate)
    max_similarities = find_max_similarities(fps)
    create_plagiarism_excel(paths.plagiarism_excel_save_path, max_similarities, plagiarism_config, excel_config)
    merge_teams_excel(paths.teams_excel_path, paths.plagiarism_excel_save_path, excel_config)
